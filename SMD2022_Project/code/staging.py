# -*- coding: utf-8 -*-
"""
Created on Sat Nov 26 17:00:46 2022

@author: arjun
"""

import numpy as np
import pandas as pd
import os
import glob
import psycopg2
from datetime import datetime
import itertools  
import openpyxl
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
class smdread(object):
    def __init__(self):
    # assign path
        return None
    
 #-----------------------------------------Filenames Reading------------------------------------   
#The followong filename method is used to return all the file names of the csv files in the path,
#such that we can then read all these files and their data   
    def filename(self,path,extension):
        p=path
        #extension = 'csv'
        os.chdir(p)
        #here the variable result has all the file names in a list
        #result = glob.glob('*.{}'.format(extension))
        result2= glob.glob(path+'/*/*.csv')
        result = glob.glob(path+'/*/*/*.csv')
        result.extend(result2)
        return result

        """
        print(result[2])
        s=len(result[2])
        print(result[2][:s-4])"""

#--------------------------------------------Data readiing------------------------------------------    
#the dataframelist method is used to read all the csv files as dataframes and append those dataframes in a list
    def dataframelist(self,filenames):
        result = filenames
        #create empty list to store all the dataframes
        dataframes_list = []

        #using read_csv, create a list of dataframes
        for i in range(len(result)):
            temp_df = pd.read_csv(result[i])
            dataframes_list.append(temp_df)
        return dataframes_list     

#metadatalist is used to convert the dataframe from above and take only the meta data from them and read them into
#dictionary with coloumn 0 as key and column 1 as value in a key value pair
    def metadatalist(self,df_list):
        
        dataframes_list = df_list
        #create an empty list where we will store all the meta data of all files in a list    
        metaDataList = []        


        #iterate through datasets in the dataframe and create a list of dictionary
        #where each dictionary has a key value pair of meta data
        for dataset in dataframes_list:
           # display(dataset)
            #print(dataset.shape)
            meta = {}
            l=[]
            l1=[]
            l=dataset.iloc[0:15,0].tolist()
            l1=dataset.iloc[0:15,1].tolist()
            for i in range(15): 
                meta[l[i]] = l1[i]
            meta[l[11]] = dataset.iloc[11,1:3].tolist()
            a=range(20)
            b=a[::2]
            for i in b:
                meta[dataset.iloc[15,i]] = dataset.iloc[15,i+1]
            meta[dataset.iloc[16,0]]=dataset.iloc[16,1]        
            metaDataList.append(meta)

        #print(metaDataList[1])    
        return metaDataList
    
    
#datalist is used to convert the dataframe from above and take only the  data from them and read them into
#list and replace Nan values and only have data part into the list    
    def datalist(self,dataframes_list):
        datalist=[]
        for dataset in dataframes_list:
            df = dataset.iloc[26:,:] 
            df.columns = df.iloc[0]
            df = df.reindex(df.index.drop(26)).reset_index(drop=True)
            df.columns.name = None
            #df2 = df.fillna(0)
            df2 = df.dropna(axis=1)
            #df.replace(np.nan, 0)
            datalist.append(df2)
        #print(datalist[0])
        #d.fillna(0)
        return datalist
        
    def hdrfilename(self,path):
        #p = list(dict.fromkeys(p))
        #extension = 'hdr'
        #os.chdir(p)
        #here the variable result has all the file names in a list
        #result = glob.glob('*.{}'.format(extension))
        p2=glob.glob(path+'/*/*/*.hdr')
        p=glob.glob(path+'/*/*/*/*.hdr')
        p.extend(p2)

        return p
    
    
    def readhdrfile(self,filepath):
        hdrlineslist=[]
        for i in filepath:
            with open(i) as f:
                lines = f.readlines()
                hdrlineslist.append(lines)                    
        return hdrlineslist
    
    def readremainingpreautismfile(self,path):
        p=glob.glob(path+'/*/*/*.dat')
        p4=glob.glob(path+'/*/*/*.wl1')
        p5=glob.glob(path+'/*/*/*.wl2')
        p3=glob.glob(path+'/*/*/*/*.dat')
        p1=glob.glob(path+'/*/*/*/*.wl1')
        p2=glob.glob(path+'/*/*/*/*.wl2')
        p.extend(p1)
        p.extend(p2)
        p.extend(p3)
        p.extend(p4)
        p.extend(p5)       
        remaininglineslist=[]
        #print(p)
        for i in p:
            with open(i) as f:
                lines = f.readlines()
                #print(lines)
                remaininglineslist.append(lines) 
        return remaininglineslist,p
    
     
    def hdrmetadata(self,hdrlinelist):
        hdrmetadatadictlist=[]
        lines=hdrlinelist
        for j in lines:
            #print(j)
            hdrmetadatadict={}
            for k in j:
                #print(k)
                
                if k.__contains__('='):
                    #print(k)
                    templist = k.split('=')
                    #print(templist)                   
                    hdrmetadatadict[templist[0]]=templist[1]
                    
            hdrmetadatadictlist.append(hdrmetadatadict) 
        return hdrmetadatadictlist
 #------------------------------------------Tables creation---------------------------------------   
    def executesqlfile(self,cursor, sql_file,db):
        for line in open(sql_file):
            cursor.execute(line)
            #print(line)
            db.commit()
    def dropschema(self,cursor,db):
        cursor.execute('DROP SCHEMA IF EXISTS smdproject CASCADE')
        db.commit()
#--------------------------------------------Data Loading----------------------------------------        
    def hubsessioninsert(self,cursor,db,filename):
        hashfilelist=[]
        for i in filename:
            hashfile = hash(i)
            hashfilelist.append(hashfile)
            time = datetime.now()
            source=hash('axw263')
            s=(hashfile,time,source)
            cursor.execute('Insert into smdproject.hubsession (sequence,timestamp,source) values(%s,%s,%s)',s )
            db.commit()
        print('hubsessioninsert is done')
        return  hashfilelist 
     
    def satsessionnameinsert(self,cursor,db,hubsessionsequencehash,vmdmetadata,hdrmetadata):
        check = hubsessionsequencehash.copy()

        for sequence,i  in zip(hubsessionsequencehash,vmdmetadata):
            time = datetime.now()
            source=hash('axw263')
            name =i['ID'][7:]
            s=(sequence,time,source,name)         
            cursor.execute('Insert into smdproject.satsessionname (sequence,timestamp,source,name) values(%s,%s,%s,%s)',s )
            check.remove(sequence)
            db.commit()

        for sequence,i in zip(check,hdrmetadata):
            time = datetime.now()
            source=hash('axw263')
            name =i['FileName'].replace('"','')
            s=(sequence,time,source,name)         
            cursor.execute('Insert into smdproject.satsessionname (sequence,timestamp,source,name) values(%s,%s,%s,%s)',s )
            db.commit()
        print('satsessionnameinsert is done')
        
        return None
    def hubmetadatainsert(self,cursor,db,vmdmetadata,hdrmetadata):
        vmcount=0
        hdrcount=2000
        hubmetadatainsertsequence = []
        hubmetadatakeys=[]
        hubmetadatavalues=[]
        for i in vmdmetadata:
            for j in list(i.items()): 
                vmcount+=1
                hashfile = hash(str(j[0])+str(vmcount))  
                hubmetadatakeys.append(str(j[0]))
                hubmetadatavalues.append(str(j[1]))
                hubmetadatainsertsequence.append(hashfile)
                time = datetime.now()
                source=hash('axw263')
                s=(hashfile,time,source)
                #print('insert %s', str(j[0]))
                cursor.execute('Insert into smdproject.hubmetadata (sequence,timestamp,source) values(%s,%s,%s)',s )
                db.commit()
        for i in hdrmetadata:
            for j in list(i.items())[0: 21]:
                hdrcount+=1
                hashfile = hash(str(j[0])+str(hdrcount))  
                hubmetadatakeys.append(str(j[0]))
                hubmetadatavalues.append(str(j[1]))
                hubmetadatainsertsequence.append(hashfile)
                time = datetime.now()
                source=hash('axw263')
                s=(hashfile,time,source)
                #print('insert %s', str(j[0]))
                cursor.execute('Insert into smdproject.hubmetadata (sequence,timestamp,source) values(%s,%s,%s)',s )
                db.commit()            
        print('hubmetadatainsert is done')           
        return hubmetadatainsertsequence,hubmetadatakeys,hubmetadatavalues
    
    def satmetadatakeyvaluepairinsert(self,cursor,db,hubmetadatainsertsequence,hubmetadatakeys,hubmetadatavalues):
        for i,j,k in zip(hubmetadatainsertsequence,hubmetadatakeys,hubmetadatavalues):
            hashfile = i
            time = datetime.now()
            source=hash('axw263')
            key = j
            value = k
            s=(hashfile,time,source,key,value)
            cursor.execute('Insert into smdproject.satmetadatakeyvaluepair (sequence,timestamp,source,key,value) values(%s,%s,%s,%s,%s)',s )
            db.commit()  
        print('satmetadatakeyvaluepairinsert is done')
        return None
    
    
    def hubsubjectinsert(self,cursor,db,vmdmetadata,hdrmetadata):
        
        hubsubjectinsertsequence=[]
        uniquesubject=[]
        for i in vmdmetadata:
            uniquesubject.append(i['Name'])
        setuniquesubject = set(uniquesubject)
        listuniquesubject=list(setuniquesubject)
        listuniquesubject.append('1')
        for i in listuniquesubject:
            hashfile = hash(i)
            hubsubjectinsertsequence.append(hashfile)
            time = datetime.now()
            source=hash('axw263')
            name = i
            s=(hashfile,time,source,name)
            cursor.execute('Insert into smdproject.hubsubject (sequence,timestamp,source,name) values(%s,%s,%s,%s)',s )
            db.commit() 
                         
        print('hubsubjectinsert is done')            
        return hubsubjectinsertsequence
    
    def satsubjectageinsert(self,cursor,db,vmdmetadata,hubsubjectinsertsequence):
        check = hubsubjectinsertsequence.copy()
        uniquesubjectage=[]
        for i in vmdmetadata:
            uniquesubjectage.append(i['Age'].replace('y',''))
        setuniquesubjectage = set(uniquesubjectage)
        listuniquesubjectage=list(setuniquesubjectage)
        for sequence,i in zip(hubsubjectinsertsequence,listuniquesubjectage):
            hashfile = sequence
            time = datetime.now()
            source=hash('axw263')
            age = i
            s=(hashfile,time,source,age)
            cursor.execute('Insert into smdproject.satsubjectage (sequence,timestamp,source,age) values(%s,%s,%s,%s)',s )
            check.remove(sequence)
            db.commit()

        print('satsubjectageinsert is done')    
        
        return None
    
    def satsubjectnameinsert(self,cursor,db,vmdmetadata,hubsubjectinsertsequence):
        uniquesubject=[]
        for i in vmdmetadata:
            uniquesubject.append(i['Name'])
        setuniquesubject = set(uniquesubject)
        listuniquesubject=list(setuniquesubject)
        listuniquesubject.append('1')
        for i,j in zip(listuniquesubject,hubsubjectinsertsequence):
            hashfile = j
            time = datetime.now()
            source=hash('axw263')
            name = i
            s=(hashfile,time,source,name)
            cursor.execute('Insert into smdproject.satsubjectname (sequence,timestamp,source,name) values(%s,%s,%s,%s)',s )
            db.commit() 
                   

        print('satsubjectnameinsert is done')    
        
        return None
    
    def hubexperimentinsert(self,cursor,db,experimentlist):
        hashfilelist=[]       
        for i in experimentlist:
            hashfile = hash(i)
            hashfilelist.append(hashfile)
            time = datetime.now()
            source=hash('axw263')
            s=(hashfile,time,source)
            cursor.execute('Insert into smdproject.hubexperiment (sequence,timestamp,source) values(%s,%s,%s)',s )
            db.commit()
        print('hubexperimentinsert is done')
        return  hashfilelist 
    
    def satexperimenttitleinsert(self,cursor,db,experimentlist,hubexperimentsession):       
        for session,i in zip(hubexperimentsession,experimentlist):
            hashfile = session            
            time = datetime.now()
            source=hash('axw263')
            title = i
            s=(hashfile,time,source,title)
            cursor.execute('Insert into smdproject.satexperimenttitle (sequence,timestamp,source,title) values(%s,%s,%s,%s)',s )
            db.commit()
        print('satexperimenttitleinsert is done')
        return  None
    def satexperimentacronyminsert(self,cursor,db,hubexperimentsession):    
        acron=['VM','PA']
        for session,k in zip(hubexperimentsession,acron):
            hashfile = session            
            time = datetime.now()
            source=hash('axw263')
            acro = k         
            s=(hashfile,time,source,acro)
            cursor.execute('Insert into smdproject.satexperimentacronym (sequence,timestamp,source,acronym) values(%s,%s,%s,%s)',s )
            db.commit()
        print('satexperimentacronyminsert is done')
        return  None
    def hubtreatmentinsert(self,cursor,db,filename,vmdmetadata,hdrmetadata,hubexperimentsession):
        hashfilelist=[]
        vmtreatmentlist=['Oxy','Deoxy','MES']
        patreatmentlist=['Normal Conversation','Stressed Conversation']
        for i in vmtreatmentlist:
            hashfile = hash(i)
            hashfilelist.append(hashfile)
            time = datetime.now()
            source=hash('axw263')
            experiment = hubexperimentsession[0]
            s = (hashfile,time,source,experiment)
            cursor.execute('Insert into smdproject.hubtreatment (sequence,timestamp,source,experiment) values(%s,%s,%s,%s)',s )
            db.commit()
            
        for i in patreatmentlist:
            hashfile = hash(i)
            hashfilelist.append(hashfile)
            time = datetime.now()
            source=hash('axw263')
            experiment = hubexperimentsession[1]
            s = (hashfile,time,source,experiment)
            cursor.execute('Insert into smdproject.hubtreatment (sequence,timestamp,source,experiment) values(%s,%s,%s,%s)',s )
            db.commit()
            
        

        print('hubtreatmentinsert is done')
        return hashfilelist
    
    def hubexperimentalunitinsert(self,cursor,db,filename,vmdmetadata):
        hashfilelist=[]
        uniquesubject=[]
        for i in vmdmetadata:
            uniquesubject.append(i['Name'])
        setuniquesubject = set(uniquesubject)
        listuniquesubject=list(setuniquesubject)
        listuniquesubject.append('1')
        for i in listuniquesubject:
            hashfile = hash(i)
            hashfilelist.append(hashfile)
            time = datetime.now()
            source=hash('axw263')
            s=(hashfile,time,source)
            cursor.execute('Insert into smdproject.hubexperimentalunit (sequence,timestamp,source) values(%s,%s,%s)',s )
            db.commit() 

        print('hubexperimentalunitinsert is done')
        return hashfilelist
    def hubfactorinsert(self,cursor,db,experimenthash):
        Factorlist=['Rest','Motor Only','Visual Only','VisuoMotor Task']
        hubfactorinserthashlist=[]
        for i in Factorlist:
            if i.__contains__('VisuoMotor'):
                hashfile = hash(i)
                hubfactorinserthashlist.append(hashfile)
                time=datetime.now()
                source=hash('axw263')
                experiment=experimenthash[0]
                iscofactor=1
                s=(hashfile,time,source,experiment,iscofactor)
                cursor.execute('Insert into smdproject.hubfactor (sequence,timestamp,source,experiment,iscofactor) values(%s,%s,%s,%s,%s)',s )
                db.commit()
            else:
                hashfile = hash(i)
                hubfactorinserthashlist.append(hashfile)
                time=datetime.now()
                source=hash('axw263')
                experiment=experimenthash[0]
                iscofactor=0
                s=(hashfile,time,source,experiment,iscofactor)
                cursor.execute('Insert into smdproject.hubfactor (sequence,timestamp,source,experiment,iscofactor) values(%s,%s,%s,%s,%s)',s )
                db.commit()
        print('hubfactorinsert is done')
        return hubfactorinserthashlist
    
    def satfactornameinsert(self,cursor,db,hubfactorinserthashlist):
        Factorlist=['Rest','Motor Only','Visual Only','VisuoMotor Task']
        for i,j in zip(Factorlist,hubfactorinserthashlist):
            hashfile = j
            time=datetime.now()
            source=hash('axw263')
            name = i
            s=(hashfile,time,source,name)
            cursor.execute('Insert into smdproject.satfactorname (sequence,timestamp,source,name) values(%s,%s,%s,%s)',s )
            db.commit()

        print('satfactornameinsert is done')
        return None
    def hubgroupinsert(self,cursor,db,vmdmetadata,treatmenthash,filename):
        hubgroupinsert=[]
        uniquegroup=[]
        count=0
        f=filename.copy()
        for i in filename:
            if i.__contains__('NIRS'):
                f.remove(i)
        for i in f:
            uniquegroup.append(i.replace('VM0001_','').replace('VM0002_','').replace('VM0003_','').replace('VM0004_','').replace('VM0005_','').replace('VM0006_','').replace('VM010_','').replace('.csv',''))
        setuniquegroup=set(uniquegroup)   
        listuniquegroup=list(setuniquegroup)
        #print(listuniquegroup)
        #print(listuniquegroup)
        for i in listuniquegroup:
            if i.__contains__('Oxy'):
                hashfile = hash(i+str(count))
                count+=1
                hubgroupinsert.append(hashfile)
                time = datetime.now()
                source=hash('axw263')
                treatment = treatmenthash[0]
                s=(hashfile,time,source,treatment)
                cursor.execute('Insert into smdproject.hubgroup (sequence,timestamp,source,treatment) values(%s,%s,%s,%s)',s )
                db.commit() 
            elif i.__contains__('Deoxy'):
                hashfile = hash(i+str(count))
                count+=1
                hubgroupinsert.append(hashfile)
                time = datetime.now()
                source=hash('axw263')
                treatment = treatmenthash[1]
                s=(hashfile,time,source,treatment)
                cursor.execute('Insert into smdproject.hubgroup (sequence,timestamp,source,treatment) values(%s,%s,%s,%s)',s )
                db.commit()
            else:
                hashfile = hash(i+str(count))
                count+=1
                hubgroupinsert.append(hashfile)
                time = datetime.now()
                source=hash('axw263')
                treatment = treatmenthash[2]
                s=(hashfile,time,source,treatment)
                cursor.execute('Insert into smdproject.hubgroup (sequence,timestamp,source,treatment) values(%s,%s,%s,%s)',s )
                db.commit()
        print('hubgroupinsert is done')
        return hubgroupinsert,listuniquegroup
    
    def satgroupnameinsert(self,cursor,db,vmdmetadata,hubgroupinserthash,listuniquegroup):
        listuniquegroup2=[]
        namelist=[]
        hashlist=[]
        for i in listuniquegroup:
            listuniquegroup2.append(i)
        for i in listuniquegroup:
            listuniquegroup2.append(i)
        for i in listuniquegroup:
            listuniquegroup2.append(i)
        for i,j in zip(listuniquegroup2,hubgroupinserthash):
            hashfile = j
            hashlist.append(hashfile)
            time = datetime.now()
            source=hash('axw263')
            name = i
            #print(name)
            namelist.append(name)
            s=(hashfile,time,source,name)
            cursor.execute('Insert into smdproject.satgroupname (sequence,timestamp,source,name) values(%s,%s,%s,%s)',s )
            db.commit() 
        print('satgroupnameinsert is done')
        return hashlist,namelist
    def assignedtoinsert(self,cursor,db,vmdmetadata,hubgrouphash,hubexperimentalunithash):
        uniquesubject=[]
        for i in vmdmetadata:
            uniquesubject.append(i['Name'])
        setuniquesubject = set(uniquesubject)
        listuniquesubject=list(setuniquesubject)
        #listuniquesubject.append('1')
        count=0
        for i,k in zip(hubexperimentalunithash[0:11],listuniquesubject):
            for j in hubgrouphash:
                hashfile = hash(k+str(count))
                count+=1
                time=datetime.now()
                source=hash('axw263')
                experimentalunit=i
                g = j
                s= (hashfile,time,source,experimentalunit,g)
                cursor.execute('Insert into smdproject.assignedto (sequence,timestamp,source,experimentalunit,grop) values(%s,%s,%s,%s,%s)',s )
                db.commit() 
        print('assignedtoinsert is done')
        return None
    
    def satfactorlevelinsert(self,cursor,db,hubfactorsession):
        satfactorlevelhash=[]
        levelvalue=['NotPresent NotPresent','NotPresent Present','Present NotPresent','Present Present']
        for i,j in zip(hubfactorsession,levelvalue):
                hashfile = i
                satfactorlevelhash.append(hashfile)
                time=datetime.now()
                source=hash('axw263')
                levelvalue=str(j)
                s= (hashfile,time,source,levelvalue)
                cursor.execute('Insert into smdproject.satfactorlevel (sequence,timestamp,source,levelvalue) values(%s,%s,%s,%s)',s )
                db.commit() 
        print('satfactorlevelinsert is done')
            
        return satfactorlevelhash
    def sattreatmentfactorlevel(self,cursor,db,hubtreatmentsession,satfactorlevelhash):
        for i,j in zip(hubtreatmentsession[0:3],satfactorlevelhash[0:3]):
            hashfile = i
            time=datetime.now()
            source=hash('axw263')
            factorlevel=j
            s= (hashfile,time,source,factorlevel)
            cursor.execute('Insert into smdproject.sattreatmentfactorlevel (sequence,timestamp,source,factorlevel) values(%s,%s,%s,%s)',s )
            db.commit() 
        print('sattreatmentfactorlevel is done')
        return None
    
    def participatesininsert(self,cursor,db,hubexperimentalunithash,hubexperimenthash):
        count=0
        participatesininserthash=[]
        for i in hubexperimentalunithash[0:7]:
            hashfile = hash('participant'+str(count))
            participatesininserthash.append(hashfile)
            count+=1
            time=datetime.now()
            source=hash('axw263')
            experimentalunit = i
            experiment = hubexperimenthash[0]
            s= (hashfile,time,source,experimentalunit,experiment)
            cursor.execute('Insert into smdproject.participatesin (sequence,timestamp,source,experimentalunit,experiment) values(%s,%s,%s,%s,%s)',s )
            db.commit()
        hashfile = hash('participant'+str(count))
        participatesininserthash.append(hashfile)
        count+=1
        time=datetime.now()
        source=hash('axw263')
        experimentalunit = hubexperimentalunithash[7]
        experiment = hubexperimenthash[1]
        s= (hashfile,time,source,experimentalunit,experiment)
        cursor.execute('Insert into smdproject.participatesin (sequence,timestamp,source,experimentalunit,experiment) values(%s,%s,%s,%s,%s)',s )
        db.commit()   
        print('participatesininsert is done')        
        return participatesininserthash
    
    def satexperimentalunitidentifierinsert(self,cursor,db,vmdmetadata,participatesinhash) :
        uniquesubject=[]
        for i in vmdmetadata:
            uniquesubject.append(i['Name'])
        setuniquesubject = set(uniquesubject)
        listuniquesubject=list(setuniquesubject)
        listuniquesubject.append('1')
        for i,j in zip(participatesinhash,listuniquesubject):
            hashfile = i
            time=datetime.now()
            source=hash('axw263')
            id = j
            s= (hashfile,time,source,id)
            cursor.execute('Insert into smdproject.satexperimentalunitidentifier (sequence,timestamp,source,ID) values(%s,%s,%s,%s)',s )
            db.commit()
        print('satexperimentalunitidentifierinsert is done')    
        return None
    
    def attendssessioninsert(self,cursor,db,hubsessionhash,hubexperimentalunithash,hubgrouphash,pafilename,satgroupnamehash,satgroupnamename):
        count=100
        count1=0
        attendssessionhash=[]
        timelist=[]
        sourcelist=[]
        experimentlist=[]
        grouplist=[]
        for i in hubexperimentalunithash[0:6]:
            for j in hubgrouphash:
                hashfile=hash(str(j)+str(count))
                attendssessionhash.append(hashfile)
                count+=1
                time=datetime.now()                
                source=hash('axw263')               
                experimentalunit = i                
                group = j
                timelist.append(time)
                sourcelist.append(source)
                experimentlist.append(experimentalunit)
                grouplist.append(group)
               
        for j,k in zip(hubgrouphash,satgroupnamename):
            if k.__contains__('Viso'):
                hashfile=hash(str(j)+str(count))
                attendssessionhash.append(hashfile)
                count+=1
                time=datetime.now()
                source=hash('axw263')
                experimentalunit = hubexperimentalunithash[6]
                group = j
                timelist.append(time)
                sourcelist.append(source)
                experimentlist.append(experimentalunit)
                grouplist.append(group)
             
        for i in pafilename:
            hashfile=hash(i+str(count))
            attendssessionhash.append(hashfile)
            count+=1
            time=datetime.now()
            source=hash('axw263')
            experimentalunit = hubexperimentalunithash[7]
            timelist.append(time)
            sourcelist.append(source)
            experimentlist.append(experimentalunit)
        for i,j,k,l,m,n in itertools.zip_longest(attendssessionhash,timelist,sourcelist,experimentlist,grouplist,hubsessionhash):
            hashfile=i
            time=j
            source=k
            experimentalunit=l
            group=m
            session=n
            s=(hashfile,time,source,experimentalunit,group,session)
            count1+=1
            cursor.execute('Insert into smdproject.attendssession (sequence,timestamp,source,experimentalunit,grop,session) values(%s,%s,%s,%s,%s,%s)',s )
            db.commit()
            
        print('attendssessioninsert is done')   
        return None
    def sessionmetadatainsert(self,cursor,db,hubmetadatasequence,hubsessionsequence):
        sessionmetadatahash=[]
        timelist=[]
        sourcelist=[]
        metadatalist=[]
        hubsessionlist=[]
        count=1
        for i in hubsessionsequence[0:75]:
            for j in hubmetadatasequence[0:26]:
                hashfile=hash(str(i)+str(j)+str(count))
                count+=1
                sessionmetadatahash.append(hashfile)
                time = datetime.now()
                timelist.append(time)
                source=hash('axw263')
                sourcelist.append(source)
                metadatalist.append(j)
                hubsessionlist.append(i)
        for i in  hubsessionsequence[75:]:
            for j in hubmetadatasequence[0:21]:
                hashfile=hash(str(i)+str(j)+str(count))
                count+=1
                sessionmetadatahash.append(hashfile)
                time = datetime.now()
                timelist.append(time)
                source=hash('axw263')
                sourcelist.append(source)
                metadatalist.append(j)
                hubsessionlist.append(i)
        for i,j,k,l,m in zip(sessionmetadatahash,timelist,sourcelist,hubsessionlist,hubmetadatasequence):
            hashfile =i
            time=j
            source=k
            session=l
            metadata=m
            s=(hashfile,time,source,session,metadata)
            cursor.execute('Insert into smdproject.sessionmetadata (sequence,timestamp,source,session,metadata) values(%s,%s,%s,%s,%s)',s )
            db.commit()
        print('sessionmetadatainsert is done')
        return None
    def hubobservationinsert(self,cursor,db,hubsessionhash,remainingdatalist):
        hubobservationhash=[]
        timelist=[]
        sourcelist=[]
        collectedatsession=[]
        count=0
        for i in hubsessionhash:
            hashfile = hash('data'+ str(count))
            hubobservationhash.append(hashfile)
            count+=1
            time =datetime.now()
            timelist.append(time)
            source=hash('axww263')
            sourcelist.append(source)
            session=i
            collectedatsession.append(session)
            # s=(hashfile,time,source,collectedatsession)
            # cursor.execute('Insert into smdproject.hubobservation (sequence,timestamp,source,collectedatsession) values(%s,%s,%s,%s)',s )
            # db.commit()   
        for i in remainingdatalist:
            hashfile = hash('data'+ str(count))
            hubobservationhash.append(hashfile)
            count+=1
            time =datetime.now()
            timelist.append(time)
            source=hash('axww263')
            sourcelist.append(source)
        for i,j,k,l in itertools.zip_longest(hubobservationhash,timelist,sourcelist,collectedatsession):
            hashfile=i
            time=j
            source=k
            session=l
            s=(hashfile,time,source,session)
            cursor.execute('Insert into smdproject.hubobservation (sequence,timestamp,source,collectedatsession) values(%s,%s,%s,%s)',s )
            db.commit() 
            
            
        print('hubobservationinsert is done')
        return hubobservationhash
    def satobservationnameinsert(self,cursor,db,hubobservationhash,allfilenames):
        for i,j in zip(hubobservationhash,allfilenames):
            hashfile =i
            time =datetime.now()
            source=hash('axww263')
            name = 'data of ' + j.replace('.csv','').replace('.hdr','').replace('.dat','').replace('.wl1','').replace('.wl2','')
            s= (hashfile,time,source,name)
            cursor.execute('Insert into smdproject.satobservationname (sequence,timestamp,source,name) values(%s,%s,%s,%s)',s )
            db.commit() 
        print('satobservationnameinsert is done')
        return None
    def  satobservationvalueinsert(self,cursor,db,hubobservationhash,vmdata,hdrmeta,remainingdata):
        check = []
        check2 =[]
        check = hubobservationhash.copy()
        for i,j in zip(hubobservationhash,vmdata):
            hashfile = i
            check.remove(i)
            time=datetime.now()
            source=hash('axw263')
            value = j.values.tolist()
            l=[]
            for j in value:
                l.append(int(j[0])-1 *0.1 )
            arr = np.array(l)
            arr2 = arr
            timearray = arr2.tolist()
            s=(hashfile,time,source,value,timearray)
            cursor.execute('Insert into smdproject.satobservationvalue (sequence,timestamp,source,value,timestamps) values(%s,%s,%s,%s,%s)',s )
            db.commit()
        check2=check.copy()
        for i,j in zip(check,hdrmeta):
            hashfile = i
            check2.remove(i)
            time=datetime.now()
            source=hash('axw263')
            value = []
            value.append(j['S-D-Key'])
            value.append(j['ChanDis'])
            timearray = [str(datetime.now()),str(datetime.now()),str(datetime.now())]
            s=(hashfile,time,source,value,timearray)
            cursor.execute('Insert into smdproject.satobservationvalue (sequence,timestamp,source,value,timestamps) values(%s,%s,%s,%s,%s)',s )
            db.commit()
        for i,j in zip(check2,remainingdata):
            hashfile = i
            time=datetime.now()
            source=hash('axw263')
            value = []
            value.append(j)
            timearray = [str(datetime.now()),str(datetime.now()),str(datetime.now())]
            s=(hashfile,time,source,value,timearray)
            cursor.execute('Insert into smdproject.satobservationvalue (sequence,timestamp,source,value,timestamps) values(%s,%s,%s,%s,%s)',s )
            db.commit()
        print('satobservationvalueinsert is done')   
        return None
    
    def observationmetadatainsert(self,cursor,db,hubobservationhash,hubmetadatahash):
        count=0
        observationmetadatahash=[]
        timelist=[]
        sourcelist=[]
        observationlist=[]
        metadatalist=[]
        for i,j in zip(hubobservationhash[0:75],hubmetadatahash[0:1950:26]):
            hashfile = hash('observationmetadata' + str(count))
            observationmetadatahash.append(hashfile)
            count+=1
            time =datetime.now()
            timelist.append(time)
            source=hash('axww263')
            sourcelist.append(source)
            observation = i
            observationlist.append(observation)
            metadata=j
            metadatalist.append(metadata)
            
        for i,j in zip(hubobservationhash[75:95],hubmetadatahash[1950:2370:21]):
            hashfile = hash('observationmetadata' + str(count))
            observationmetadatahash.append(hashfile)
            count+=1
            time =datetime.now()
            timelist.append(time)
            source=hash('axww263')
            sourcelist.append(source)
            observation = i
            observationlist.append(observation)
            metadata=j
            metadatalist.append(metadata)
            
        for i in hubobservationhash[95:]:
            hashfile = hash('observationmetadata' + str(count))
            observationmetadatahash.append(hashfile)
            count+=1
            time =datetime.now()
            timelist.append(time)
            source=hash('axww263')
            sourcelist.append(source)
            observation = i
            observationlist.append(observation)
            
            
        for i,j,k,l,m in itertools.zip_longest(observationmetadatahash,timelist,sourcelist,observationlist,metadatalist):
            hashfile=i
            time=j
            source=k
            observation=l
            metadata=m
            s= (hashfile,time,source,observation,metadata)
            cursor.execute('Insert into smdproject.observationmetadata (sequence,timestamp,source,observation,metadata) values(%s,%s,%s,%s,%s)',s )
            db.commit()
        print('observationmetadatainsert is done')
            
        return None
    
    
#----------------------------------------------Execution----------------------------------------------    
class Test:
    def main():
        #p ="C:/Users/arjun/Documents/smd project/Dataset1_VM_BlindedAndReduced/VMData_Blinded/"
        print('Tables creation is starting')
        a = os.path.abspath(os.getcwd())
        p=str(a)+'/'
        #print(p)
        r = smdread()        
        extension='csv'        
        vmreaderfilename = r.filename(p,extension)
        df = r.dataframelist(vmreaderfilename)
        mdl=r.metadatalist(df)
        dl = r.datalist(df)
        #print(mdl)
        #p2="C:/Users/arjun/Documents/smd project/Dataset2_PreAutism_BlindedAndReduced/PreAutismData_Blinded/"        
        p2=str(a)+'/'
        r2= smdread()
        preautismfilename= r2.hdrfilename(p2)
        hdrlist=r2.readhdrfile(preautismfilename)
        remain,remainingfilepaths=r2.readremainingpreautismfile(p2)
        #print(hdrlist)
        hdrmeta = r2.hdrmetadata(hdrlist)
        
        vmfilenames=[]
        for i in vmreaderfilename:
            j=i.split('\\')
            
            vmfilenames.append(j[-1])

        #allfilenames= vmreaderfilename
        allfilenames= vmfilenames
        preaustism = []
        for i in preautismfilename:
            preaustism.append(i[-23:])
            
        #print(preaustism)    
        allfilenames.extend(preaustism)
        #print(allfilenames)
        allfilenames2=[]
        allfilenames2.extend(allfilenames)
        
        #allfilenames2.extend(remainingfilepaths)
        configpath =glob.glob(p+'config.txt')
        configfile=str(configpath[0])
        #print(configfile)
        with open(configfile) as f:
            config = f.readlines()
        #print(config)
        configuration1=[]
        configuration=[]
        for i in config:
            configuration1.extend(i.split('='))
        for i in configuration1:
            configuration.append(i.replace(',','').replace('"','').replace('\n','').replace(';','').replace(':','').replace(' ','').replace('\t',''))
        #print(configuration1)
        #print(configuration)
        Db=configuration[1]
        ht=configuration[3]
        us=configuration[5]
        pwd=configuration[7]
        pt=configuration[9]
            
                
        #print(Db,ht,us,pwd,pt)       
        for i in remainingfilepaths:
            allfilenames2.append(i[-23:])
            
            
       
        conn = psycopg2.connect(database=Db,
                        host=ht,
                        user=us,
                        password=pwd,
                        port=pt)
        cursor = conn.cursor()
        conn.autocommit = True       
        cursor = conn.cursor()
        
        r2.dropschema(cursor,conn)
        a = os.path.abspath(os.getcwd())
        p=str(a)+'/'
        sqlpath = glob.glob(p+'/staging.sql')
        sqlfile = str(sqlpath[0])                  

        r2.executesqlfile(cursor,sqlfile,conn)
        experimentlist=['VM_BlindedAndReduced','PreAutism_BlindedAndReduced']
        conn.commit()
        print('All Tables have been created')
        print('Data Insertion is starting')
        hubsesion=r2.hubsessioninsert(cursor,conn,allfilenames)
        hubexperiment = r2.hubexperimentinsert(cursor,conn,experimentlist)
        r2.satexperimenttitleinsert(cursor,conn,experimentlist,hubexperiment)
        r2.satexperimentacronyminsert(cursor,conn,hubexperiment)
        hubtreatment = r2.hubtreatmentinsert(cursor,conn,allfilenames,mdl,hdrmeta,hubexperiment)
        r2.satsessionnameinsert(cursor,conn,hubsesion,mdl,hdrmeta)
        hubmetadatainsertsequence,hubmetadatakeys,hubmetadatavalues = r2.hubmetadatainsert(cursor, conn,mdl,hdrmeta)
        r2.satmetadatakeyvaluepairinsert(cursor,conn,hubmetadatainsertsequence,hubmetadatakeys,hubmetadatavalues)
        hubsubject = r2.hubsubjectinsert(cursor,conn,mdl,hdrmeta)
        r2.satsubjectageinsert(cursor,conn,mdl,hubsubject)
        r2.satsubjectnameinsert(cursor,conn,mdl,hubsubject)
        hubexperimentalunit = r2.hubexperimentalunitinsert(cursor,conn,allfilenames,mdl)
        hubfactor=r2.hubfactorinsert(cursor,conn,hubexperiment)
        r2.satfactornameinsert(cursor,conn,hubfactor)
        hubgroup,listuniquegroup = r2.hubgroupinsert(cursor,conn,mdl,hubtreatment,allfilenames)
        satgroupnamehash,satgroupnamename = r2.satgroupnameinsert(cursor, conn, mdl, hubgroup,listuniquegroup)
        r2.assignedtoinsert(cursor,conn,mdl,hubgroup,hubexperimentalunit)
        satfactorlevel=r2.satfactorlevelinsert(cursor,conn,hubfactor)
        r2.sattreatmentfactorlevel(cursor,conn,hubtreatment,satfactorlevel)
        participatesin = r2.participatesininsert(cursor,conn,hubexperimentalunit,hubexperiment)
        r2.satexperimentalunitidentifierinsert(cursor,conn,mdl,participatesin)
        r2.attendssessioninsert(cursor,conn,hubsesion,hubexperimentalunit,hubgroup,preaustism,satgroupnamehash,satgroupnamename)
        r2.sessionmetadatainsert(cursor,conn,hubmetadatainsertsequence,hubsesion)
        hubobservation = r2.hubobservationinsert(cursor,conn,hubsesion,remain)
        r2.satobservationnameinsert(cursor,conn,hubobservation,allfilenames2)
        r2.observationmetadatainsert(cursor,conn,hubobservation,hubmetadatainsertsequence)
        r2.satobservationvalueinsert(cursor,conn,hubobservation,dl,hdrmeta,remain)
        print('Data insertion is complete for all the Tables')
        print('Information layer querying is starting')
        
        p=str(a)+'/'
        sqlpath = glob.glob(p+'/Informationlayerqueries.sql')
        sqlfile = str(sqlpath[0]) 
        for line in open(sqlfile):
            cursor.execute(line)
            records = cursor.fetchall() 
            count =3
            pd.DataFrame(records).to_excel('Information Layer query results.xlsx', sheet_name= 'Query ' + str(count)+' results')
            break;
        count=3
        for line in open(sqlfile):
            cursor.execute(line)
            records = cursor.fetchall() 
            
            with pd.ExcelWriter('Information Layer query results.xlsx', engine="openpyxl", mode="a",if_sheet_exists="replace") as writer:
                pd.DataFrame(records).to_excel(writer, sheet_name= 'Query ' + str(count)+' results')
                count+=1

        print('Information layer querying is completed,open Information Layer query results.xlsx file for results')
        
       
        
        
 #-----------------------------------------Information Querying----------------------------------       
        for i in dl:
            plt.plot(i['Time'][0:10],dl[0]['CH1'][0:10])
            plt.xlabel('Time')
            #plt.yticks(color='w')
            plt.tick_params(
            axis='x',          # changes apply to the x-axis
            which='both',      # both major and minor ticks are affected
            bottom=False,      # ticks along the bottom edge are off
            top=False,         # ticks along the top edge are off
            labelbottom=False)
            plt.ylabel('Y label')
        plt.savefig('Plot1.png')
        plt.show()  
       
        a=dl[0].to_numpy()
        floata=np.asarray(a,dtype=float)
        plt.boxplot(floata)
        plt.xlabel('boxplot x')
        plt.ylabel('boxplot y')
        plt.savefig('Plot2.png')
        plt.show() 
        
       
        wb = openpyxl.load_workbook('Information Layer query results.xlsx')
        wb.create_sheet('Query1 Plot')
        active = wb['Query1 Plot']
        active.add_image(Image('Plot1.png'),'H1')
        wb.create_sheet('Query6 Plot')
        active = wb['Query6 Plot']
        active.add_image(Image('Plot2.png'),'H1')
        wb.save('Information Layer query results.xlsx')
        
        
         
        # show plot
        plt.show()
        
        print('Code execution is complete')
                
       
    
    
if __name__ == "__main__":
    ##COMPLETE THIS PART
    Test.main()